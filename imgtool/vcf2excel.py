#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VCF 通讯录文件转 Excel 工具
支持解析标准 vCard 格式 (.vcf) 并导出为 Excel (.xlsx)
"""

import os
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def normalize_tel_type(tel_type):
    """标准化电话类型"""
    if not tel_type:
        return '手机'

    type_map = {
        'CELL': '手机', 'WORK': '工作', 'HOME': '家庭',
        'VOICE': '语音', 'FAX': '传真', 'MAIN': '主要',
        'PREF': '偏好', 'OTHER': '其他', 'IPHONE': '手机',
        '住宅': '家庭', '办': '工作', '手': '手机', '家': '家庭',
        '手机': '手机', '工作': '工作', '家庭': '家庭', '传真': '传真',
    }

    # 直接匹配
    if tel_type in type_map:
        return type_map[tel_type]

    # 包含关键词匹配
    if '移动' in tel_type or '联通' in tel_type or '电信' in tel_type:
        return '手机'
    elif '住宅' in tel_type or '家庭' in tel_type:
        return '家庭'
    elif '办公' in tel_type or '公司' in tel_type or '单位' in tel_type:
        return '工作'
    elif '传真' in tel_type:
        return '传真'
    elif '主要' in tel_type:
        return '主要'
    elif '偏好' in tel_type:
        return '偏好'

    # 包含地名的归为手机
    cities = ['北京', '上海', '广州', '深圳', '成都', '杭州', '南京', '武汉', '西安', '长沙']
    if any(city in tel_type for city in cities):
        return '手机'

    return '手机'


def parse_vcf_file(file_path):
    """
    解析 VCF 文件，返回联系人列表
    """
    contacts = []

    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 按 BEGIN:VCARD 分割联系人
    vcard_blocks = re.split(r'(?=BEGIN:VCARD)', content)

    for block in vcard_blocks:
        if not block.strip() or 'BEGIN:VCARD' not in block:
            continue

        contact = {'姓名': '', '电话列表': [], '邮箱列表': []}
        lines = block.strip().split('\n')

        # 折叠续行处理
        unfolded_lines = []
        for line in lines:
            if line.startswith((' ', '\t')) and unfolded_lines:
                unfolded_lines[-1] += line[1:]
            else:
                unfolded_lines.append(line)

        # 收集 item 标签映射
        item_labels = {}
        for line in unfolded_lines:
            if ':' not in line:
                continue
            colon_idx = line.index(':')
            field = line[:colon_idx]
            value = line[colon_idx + 1:].strip()

            match = re.match(r'item(\d+)\.x-ablabel$', field, re.IGNORECASE)
            if match:
                item_num = match.group(1)
                item_labels[item_num] = value.strip()

        # 处理字段
        for line in unfolded_lines:
            if ':' not in line:
                continue

            colon_idx = line.index(':')
            field = line[:colon_idx]
            value = line[colon_idx + 1:].strip()

            # 跳过 X-ABLabel
            if re.match(r'item\d+\.x-ablabel$', field, re.IGNORECASE):
                continue

            # 处理 ITEM 前缀
            item_num = None
            item_match = re.match(r'item(\d+)\.(.+)$', field, re.IGNORECASE)
            if item_match:
                item_num = item_match.group(1)
                field = item_match.group(2)

            # 解析属性名和参数
            field_parts = field.split(';')
            field_name = field_parts[0].upper()
            field_params = {}

            for part in field_parts[1:]:
                if '=' in part:
                    key, val = part.split('=', 1)
                    field_params[key.upper()] = val

            # 获取自定义标签
            custom_label = item_labels.get(item_num) if item_num else None

            if field_name in ('FN', 'N'):
                if field_name == 'N' and contact['姓名']:
                    continue
                contact['姓名'] = value
            elif field_name == 'TEL':
                tel_type = normalize_tel_type(custom_label)
                if not custom_label:
                    type_val = field_params.get('TYPE', '')
                    if isinstance(type_val, list):
                        for t in type_val:
                            if t.upper() not in ['CELL', 'VOICE', 'PREF', 'VOICE']:
                                tel_type = normalize_tel_type(t)
                                break
                    elif type_val:
                        tel_type = normalize_tel_type(type_val)

                contact['电话列表'].append({'type': tel_type, 'value': value})
            elif field_name == 'EMAIL':
                email_type = custom_label if custom_label else ''
                if not email_type:
                    type_val = field_params.get('TYPE', '')
                    if isinstance(type_val, list):
                        type_val = type_val[0] if type_val else ''
                    if type_val and type_val.upper() != 'INTERNET':
                        email_type = type_val

                contact['邮箱列表'].append({'type': email_type, 'value': value})
            elif field_name == 'ORG':
                contact['公司'] = value
            elif field_name == 'TITLE':
                contact['职位'] = value
            elif field_name == 'ADR':
                contact['地址'] = value
            elif field_name == 'URL':
                contact['网址'] = value
            elif field_name == 'NOTE':
                contact['备注'] = value
            elif field_name == 'BDAY':
                contact['生日'] = value
            elif field_name == 'NICKNAME':
                contact['昵称'] = value

        # 整理电话列表到不同列
        tel_priority = ['手机', '工作', '家庭', '主要', '偏好', '传真', '其他', '语音',
                        '备用 1', '备用 2', '备用 3', '备用 4', '备用 5',
                        '备用 6', '备用 7', '备用 8', '备用 9', '备用 10']
        tel_columns = {t: None for t in tel_priority}
        extra_phones = []

        for tel in contact['电话列表']:
            t = tel['type']
            v = tel['value']
            placed = False
            if t in tel_columns:
                if tel_columns[t] is None:
                    tel_columns[t] = v
                    placed = True

            if not placed:
                # 尝试放入备用列
                for fallback in tel_priority:
                    if tel_columns[fallback] is None:
                        tel_columns[fallback] = v
                        placed = True
                        break

            if not placed:
                # 所有列都满了，存入额外列表
                extra_phones.append(v)

        for t, v in tel_columns.items():
            if v is not None:
                contact[f'电话 ({t})'] = v

        # 如果还有多余的电话，合并到'备用 10'或追加
        if extra_phones:
            if contact.get('电话 (备用 10)'):
                contact['电话 (备用 10)'] = contact['电话 (备用 10)'] + '; ' + '; '.join(extra_phones)
            else:
                contact['电话 (备用 10)'] = '; '.join(extra_phones)

        # 整理邮箱列表
        email_priority = ['', '工作', '个人', '偏好']
        email_columns = {t: None for t in email_priority}

        for email in contact['邮箱列表']:
            t = email['type']
            v = email['value']
            if t in email_columns:
                if email_columns[t] is None:
                    email_columns[t] = v
                else:
                    for fallback in email_priority:
                        if email_columns[fallback] is None:
                            email_columns[fallback] = v
                            break
            else:
                email_columns[t] = v

        for t, v in email_columns.items():
            if v is not None:
                key = f'邮箱 ({t})' if t else '邮箱'
                contact[key] = v

        # 删除临时字段
        del contact['电话列表']
        del contact['邮箱列表']

        if contact.get('姓名'):
            contacts.append(contact)

    return contacts


def collect_all_fields(contacts):
    """收集所有字段名"""
    all_fields = set()
    for contact in contacts:
        all_fields.update(contact.keys())

    priority_fields = [
        '姓名', '电话 (手机)', '电话 (工作)', '电话 (家庭)', '电话 (主要)',
        '电话 (偏好)', '电话 (传真)', '电话 (其他)', '电话 (语音)',
        '电话 (备用 1)', '电话 (备用 2)', '电话 (备用 3)', '电话 (备用 4)',
        '电话 (备用 5)', '电话 (备用 6)', '电话 (备用 7)', '电话 (备用 8)',
        '电话 (备用 9)', '电话 (备用 10)',
        '邮箱', '邮箱 (工作)', '邮箱 (个人)', '邮箱 (偏好)',
        '公司', '职位', '昵称', '地址', '网址', '生日', '备注'
    ]

    ordered_fields = []
    for field in priority_fields:
        if field in all_fields:
            ordered_fields.append(field)
            all_fields.remove(field)

    # 过滤不需要的字段
    unwanted = {'VERSION', 'BEGIN', 'END', 'PRODID', 'PHOTO', 'IMPP', 'X_AIM',
                'X_PHONETIC_FIRST_NAME', 'X_PHONETIC_LAST_NAME', 'X_ABLABEL', 'N', 'FN'}
    all_fields -= unwanted

    # 过滤已标准化的电话/邮箱字段
    filtered = {f for f in all_fields if not f.startswith('电话 (') and not f.startswith('邮箱')}

    ordered_fields.extend(sorted(filtered))
    return ordered_fields


def export_to_excel(contacts, output_path):
    """导出为 Excel"""
    if not contacts:
        print("没有可导出的联系人数据")
        return

    all_fields = collect_all_fields(contacts)

    wb = Workbook()
    ws = wb.active
    ws.title = "通讯录"

    # 表头
    for col, field in enumerate(all_fields, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 数据
    for row_idx, contact in enumerate(contacts, 2):
        for col_idx, field in enumerate(all_fields, 1):
            value = contact.get(field, '')
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column].width = adjusted_width

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"成功导出 {len(contacts)} 条联系人记录到：{output_path}")


def verify_data(vcf_path, xlsx_path):
    """验证数据准确性"""
    print("\n=== 数据验证 ===")

    # 统计 VCF
    with open(vcf_path, 'r', encoding='utf-8') as f:
        vcf_content = f.read()
    vcf_count = vcf_content.count('BEGIN:VCARD')

    # 统计 Excel
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    excel_count = ws.max_row - 1

    print(f"VCF 联系人数量：{vcf_count}")
    print(f"Excel 联系人数量：{excel_count}")
    print(f"数量匹配：{'OK' if vcf_count == excel_count else 'FAIL'}")

    # 抽样验证电话号码
    print("\n=== 抽样验证 ===")
    contacts = parse_vcf_file(vcf_path)

    # 验证有分号分隔电话的联系人
    sample_count = 0
    for contact in contacts[:10]:
        name = contact.get('姓名', '')
        tels = [v for k, v in contact.items() if k.startswith('电话 (') and v]
        if len(tels) > 1:
            print(f"{name}: {tels}")
            sample_count += 1
            if sample_count >= 5:
                break

    return vcf_count == excel_count


def main():
    parser = argparse.ArgumentParser(
        description='VCF 通讯录文件转 Excel 工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例用法:
  python vcf2excel.py contacts.vcf
  python vcf2excel.py contacts.vcf -o output.xlsx
  python vcf2excel.py input.vcf --output my_contacts.xlsx
        '''
    )

    parser.add_argument('input', help='输入的 VCF 文件路径')
    parser.add_argument('-o', '--output', help='输出的 Excel 文件路径 (默认：输入文件名.xlsx)')
    parser.add_argument('-v', '--verify', action='store_true', help='转换后验证数据')

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误：文件不存在 - {args.input}")
        return 1

    if args.output:
        output_path = args.output
    else:
        base_name = os.path.splitext(args.input)[0]
        output_path = base_name + '.xlsx'

    print(f"正在解析 VCF 文件：{args.input}")
    contacts = parse_vcf_file(args.input)

    if not contacts:
        print("未找到任何联系人数据")
        return 1

    print(f"找到 {len(contacts)} 条联系人记录")

    export_to_excel(contacts, output_path)

    if args.verify:
        verify_data(args.input, output_path)

    return 0


if __name__ == '__main__':
    exit(main())
